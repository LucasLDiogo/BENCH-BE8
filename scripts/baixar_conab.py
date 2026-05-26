"""
Coletor CONAB — Acompanhamento da Safra Brasileira de Grãos.
Fontes possíveis (em ordem de preferência):
  1) Série histórica consolidada (XLSX): portaldeinformacoes.conab.gov.br
  2) Tabela de dados do boletim mensal vigente (XLS/XLSX em /info-agro/safras/graos)
  3) Tabela auxiliar disponível em dados.gov.br

Saída: data/conab_graos.json com Soja, Milho (total) e Trigo, por UF.
Se nenhuma fonte responder, salva JSON com status ERRO e estrutura vazia
(o HTML mostrará "Fonte indisponível" sem quebrar).
"""
from __future__ import annotations
import re, datetime as dt
from utils import (http_get, save_json, mark_source_ok, mark_source_error, mark_source_partial,
                   DATA_DIR, DOWNLOADS_DIR, log, now_iso, pct)

CONAB_LANDING = "https://www.conab.gov.br/info-agro/safras/graos"
CONAB_SERIE_HIST = "https://portaldeinformacoes.conab.gov.br/safra-serie-historica-graos.html"

# URLs candidatas do XLSX mais recente (padrão CONAB)
def _candidate_urls() -> list[str]:
    return [
        "https://www.conab.gov.br/info-agro/safras/graos/boletim-da-safra-de-graos/SerieHistoricaGraos.xlsx",
        "https://www.conab.gov.br/info-agro/safras/graos/boletim-da-safra-de-graos/serie-historica-de-graos",
        "https://portaldeinformacoes.conab.gov.br/downloads/arquivos/SerieHistoricaGraos.xlsx",
        "https://portaldeinformacoes.conab.gov.br/downloads/arquivos/SerieHistoricaGraosAgosto2025.xlsx",
        "https://portaldeinformacoes.conab.gov.br/downloads/arquivos/SerieHistoricaGraos2025.xlsx",
    ]

def discover_xlsx_link() -> str | None:
    """Procura link do XLSX da Série Histórica na página do boletim."""
    for landing in [CONAB_LANDING + "/boletim-da-safra-de-graos", CONAB_LANDING, CONAB_SERIE_HIST]:
        try:
            html = http_get(landing, timeout=20).decode("utf-8", errors="ignore")
            # Procura .xlsx / .xls
            links = re.findall(r'href="([^"]+\.(?:xlsx|xls))"', html, re.IGNORECASE)
            # Prefere "SerieHistoricaGraos"
            for L in links:
                if "seriehistorica" in L.lower() or "serie_historica" in L.lower():
                    if L.startswith("/"):
                        L = "https://www.conab.gov.br" + L
                    return L
            if links:
                L = links[0]
                if L.startswith("/"):
                    L = "https://www.conab.gov.br" + L
                return L
        except Exception as e:
            log.debug(f"CONAB landing {landing}: {e}")
            continue
    return None

def _try_download() -> tuple[bytes | None, str | None]:
    url = discover_xlsx_link()
    if url:
        try:
            data = http_get(url, timeout=60)
            if data and len(data) > 10000:
                log.info(f"CONAB · XLSX descoberto: {url}")
                return data, url
        except Exception as e:
            log.warning(f"CONAB descoberta falhou: {e}")
    for u in _candidate_urls():
        try:
            log.info(f"CONAB · tentando fallback {u}")
            data = http_get(u, timeout=60)
            if data and len(data) > 10000:
                return data, u
        except Exception:
            continue
    return None, None

def parse_serie_historica_xlsx(raw: bytes) -> dict:
    """Parser estrutural da planilha CONAB · Série Histórica de Grãos.
    A planilha tem múltiplas abas (Soja, Milho 1ª safra, Milho 2ª safra, Trigo, etc.)
    e cada aba tem colunas: UF | Produção | Área | Produtividade × safra.
    """
    try:
        import openpyxl
        from io import BytesIO
        wb = openpyxl.load_workbook(BytesIO(raw), data_only=True)
    except ImportError:
        raise RuntimeError("openpyxl não instalado")
    except Exception as e:
        raise RuntimeError(f"abertura falhou: {e}")

    log.info(f"CONAB · abas encontradas: {wb.sheetnames[:12]}")

    culturas = {}
    # Procura abas relevantes
    for sheet_name in wb.sheetnames:
        low = sheet_name.lower()
        cultura = None
        if "soja" in low and "total" not in low: cultura = "Soja"
        elif "milho" in low and ("total" in low or " " not in low.replace("milho","").strip()): cultura = "Milho (total)"
        elif "trigo" in low: cultura = "Trigo"
        if not cultura: continue

        ws = wb[sheet_name]
        rows = list(ws.iter_rows(values_only=True))
        # Detectar safra atual (última coluna preenchida no header)
        # CONAB padrão: header em linhas iniciais com "Safra"
        header_idx = None
        for i, r in enumerate(rows[:20]):
            if r and any(isinstance(c, str) and ("safra" in (c or "").lower() or "uf" in (c or "").lower()) for c in r if c):
                header_idx = i
                break
        if header_idx is None: continue

        # Extrai linhas com UF + número
        registros = []
        for r in rows[header_idx+1:]:
            if not r or not r[0]: continue
            uf = str(r[0]).strip().upper()
            if len(uf) != 2 or not uf.isalpha(): continue
            # Pega últimos números à direita (Produção / Área)
            nums = [c for c in r[1:] if isinstance(c, (int, float))]
            if not nums: continue
            registros.append({
                "uf": uf,
                "producao_mil_t": float(nums[-3]) if len(nums) >= 3 else None,
                "area_mil_ha":    float(nums[-2]) if len(nums) >= 2 else None,
                "produtividade_kg_ha": float(nums[-1]) if len(nums) >= 1 else None,
            })
        if registros:
            culturas[cultura] = registros
            log.info(f"CONAB · {cultura}: {len(registros)} UFs · {sheet_name}")

    return culturas

def consolidate(culturas: dict) -> dict:
    """Calcula totais e shares por cultura."""
    out = {"culturas": []}
    safra_atual = f"{dt.date.today().year-1}/{str(dt.date.today().year)[2:]}"
    for nome, regs in culturas.items():
        total_prod = sum(r["producao_mil_t"] or 0 for r in regs)
        total_area = sum(r["area_mil_ha"] or 0 for r in regs)
        # Calcula share
        ufs_ranked = []
        for r in sorted(regs, key=lambda x: x["producao_mil_t"] or 0, reverse=True):
            share = (r["producao_mil_t"] / total_prod * 100) if (total_prod and r["producao_mil_t"]) else None
            ufs_ranked.append({
                "uf": r["uf"],
                "producao_mil_t": r["producao_mil_t"],
                "area_mil_ha": r["area_mil_ha"],
                "produtividade_kg_ha": r["produtividade_kg_ha"],
                "share_pct": round(share, 2) if share else None,
            })
        out["culturas"].append({
            "cultura": nome,
            "safra": safra_atual,
            "producao_total_mt":  round(total_prod / 1000, 2) if total_prod else None,  # mil t → Mt
            "area_total_mha":     round(total_area / 1000, 2) if total_area else None,
            "produtividade_media_kg_ha": round(
                sum((r["produtividade_kg_ha"] or 0)*(r["producao_mil_t"] or 0) for r in regs) /
                (sum(r["producao_mil_t"] or 0 for r in regs) or 1), 0
            ) if total_prod else None,
            "ufs": ufs_ranked,
            "top_uf": ufs_ranked[0]["uf"] if ufs_ranked else None,
        })
    return out

def run() -> None:
    out = {
        "ultima_atualizacao": now_iso(),
        "fonte": "CONAB · Acompanhamento da Safra Brasileira de Grãos",
        "endpoint": CONAB_LANDING,
        "status": "OK",
        "safra_referencia": None,
        "culturas": [],
    }
    raw, url_used = _try_download()
    if not raw:
        log.error("CONAB: planilha indisponível em todas as URLs candidatas")
        out["status"] = "ERRO"
        out["erro"] = "Planilha CONAB indisponível"
        out["nota"] = "Fonte indisponível — execução tentou descoberta + 5 URLs candidatas"
        save_json(DATA_DIR / "conab_graos.json", out)
        mark_source_error("conab_graos", "Planilha indisponível", endpoint=CONAB_LANDING)
        return

    # Salva raw
    fname = f"SerieHistoricaGraos_{dt.date.today().isoformat()}.xlsx"
    (DOWNLOADS_DIR / "conab" / fname).write_bytes(raw)
    log.info(f"CONAB · raw salvo: {fname}")

    try:
        culturas = parse_serie_historica_xlsx(raw)
        if not culturas:
            raise RuntimeError("nenhuma aba relevante encontrada (Soja/Milho/Trigo)")
        cons = consolidate(culturas)
        out.update(cons)
        out["safra_referencia"] = cons["culturas"][0]["safra"] if cons["culturas"] else None
        out["url_xlsx"] = url_used
        save_json(DATA_DIR / "conab_graos.json", out)
        n_ufs = sum(len(c["ufs"]) for c in cons["culturas"])
        mark_source_ok("conab_graos", rows=n_ufs,
                       note=f"{len(cons['culturas'])} culturas · {n_ufs} registros UF",
                       endpoint=url_used)
        log.info(f"CONAB OK · {len(cons['culturas'])} culturas · {n_ufs} registros")
    except RuntimeError as e:
        if "openpyxl" in str(e):
            out["status"] = "PARCIAL"
            out["nota"] = "Planilha CONAB capturada, parser openpyxl indisponível neste ambiente"
            save_json(DATA_DIR / "conab_graos.json", out)
            mark_source_partial("conab_graos", "planilha capturada · parser pendente",
                                rows=0, endpoint=url_used)
        else:
            log.exception("CONAB · parser falhou")
            out["status"] = "ERRO"
            out["erro"] = str(e)[:300]
            save_json(DATA_DIR / "conab_graos.json", out)
            mark_source_error("conab_graos", str(e), endpoint=url_used)

if __name__ == "__main__":
    run()
